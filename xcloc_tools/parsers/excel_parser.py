import openpyxl
import pathlib
from models import Contents, Tool, LocalizedContents, Translation

class ExcelParser:

    def __init__(self, path):
        if path.exists() == False:
            raise IOError("No such file or directory: {}".format(path))

        self.file_path = path

    def parse(self):
        contents = Contents(file = str(self.file_path))

        wb = openpyxl.load_workbook(self.file_path)

        self.__parse_metadata_sheet(contents, wb)
        self.__parse_translation_sheets(contents, wb)

        return contents

    def __parse_metadata_sheet(self, contents, wb):
        index = wb["_metadata_"]

        contents.file = pathlib.Path(index["C2"].value)
        contents.version = index["B2"].value
        contents.source = index["D2"].value
        contents.target = index["E2"].value

        tool = Tool()
        tool.id = index["F2"].value
        tool.name = index["G2"].value
        tool.version = index["H2"].value
        tool.build = index["I2"].value
        contents.tool = tool


    def __parse_translation_sheets(self, contents, wb):
        # Get metadata
        index = wb["_metadata_"]
        temp_localized_contents = []
        for i in range(4, index.max_row + 1):
            localized_contents = LocalizedContents()

            localized_contents.file = index["C" + str(i)].value
            localized_contents.source = index["D" + str(i)].value
            localized_contents.target = index["E" + str(i)].value

            tool = Tool()
            tool.id = index["F" + str(i)].value
            tool.name = index["G" + str(i)].value
            tool.version = index["H" + str(i)].value
            tool.build = index["I" + str(i)].value
            localized_contents.tool = tool

            temp_localized_contents.append(localized_contents)

        # Parse all sheets
        for sheetname in wb.sheetnames:
            if sheetname == "_metadata_":
                continue

            ws = wb[sheetname]
            row = 1

            if ws["A" + str(row)].value != "No":
                row += 1
            
            if ws["A" + str(row)].value != "No":
                print("Sheet format is invalidated : {}".format(sheetname))
                continue

            row += 1

            # Get localized content with metadata
            filtered_localized_contents = list(filter(lambda c: c.basename == sheetname, temp_localized_contents))
            if filtered_localized_contents == False:
                print("Sheet metadata not found: {}".format(sheetname))
                continue

            localized_contents = filtered_localized_contents[0]

            # Add translations
            for i in range(row, ws.max_row + 1):
                translation = Translation()
                translation.id = ws["B" + str(i)].value
                translation.source = ws["C" + str(i)].value
                translation.target = ws["D" + str(i)].value
                translation.note = ws["E" + str(i)].value

                localized_contents.translations.append(translation)
            
            contents.localized_contents.append(localized_contents)
