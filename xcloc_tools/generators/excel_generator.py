import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
from models import Contents, Tool, LocalizedContents, Translation

class ExcelGenerator:

    def __init__(self, contents):
        self.contents = contents

    def generate(self):
        wb = openpyxl.Workbook()

        self.__generate_metadata_sheet(wb)
        self.__generate_translation_sheets(wb)

        wb.save((output := '{}.xlsx'.format(self.contents.stem)))

        return output

    def __generate_metadata_sheet(self, wb):
        ws = wb.active
        ws.title = "_metadata_"
        ws.append(["Source File", "Version","Path", "Region", "Target", "Tool", "Name", "Version", "Build"])
        ws.append([ self.contents.basename, self.contents.version, self.contents.file, self.contents.source, self.contents.target, self.contents.tool.id, self.contents.tool.name, self.contents.tool.version, self.contents.tool.build])
        ws.append([])

        ws.append(["File", "Items", "Path", "Source", "Target", "Tool", "Name", "Version", "Build"])
        for file in self.contents.localized_contents:
             ws.append([file.basename, len(file.translations), file.file, file.source, file.target, file.tool.id, file.tool.name, file.tool.version, file.tool.build])

        # Column width
        ws.column_dimensions["A"].width = 49.17
        ws.column_dimensions["B"].width = 8.00
        ws.column_dimensions["C"].width = 15.83
        ws.column_dimensions["D"].width = 15.83
        ws.column_dimensions["E"].width = 15.83
        ws.column_dimensions["F"].width = 15.83
        ws.column_dimensions["G"].width = 15.83
        ws.column_dimensions["H"].width = 15.83
        ws.column_dimensions["I"].width = 15.83

        # Borders
        for i in range(1, 10):
            ws.cell(row = 1, column = i).border = Border(bottom = Side(style = "thin"))
            ws.cell(row = 1, column = i).font = Font(bold = True)

        for i in range(1, 10):
            ws.cell(row = 4, column = i).border = Border(bottom = Side(style="thin"))
            ws.cell(row = 4, column = i).font = Font(bold = True)

    def __generate_translation_sheets(self, wb):
        for file in self.contents.localized_contents:
            ws = wb.create_sheet(file.basename)

            row = 1

            # Add screenshot if exists
            if file.basename in self.contents.screenshots and (screenshot := self.contents.screenshots[file.basename]):
                # Cell height
                ws.row_dimensions[row].height = 400

                img = openpyxl.drawing.image.Image(screenshot)
                if img.height >= 480:
                    img.width = img.width * (480 / img.height)
                    img.height = 480
                img.anchor = 'B' + str(row)
                ws.add_image(img)
                ws.append([])

                row += 1

            ws.append(["No", "ID", "Source", "Target", "Note"])

            for no, translation in enumerate(file.translations, start = 1):
                ws.append([no, translation.id, translation.source, translation.target, translation.note])

            # Column width
            ws.column_dimensions["A"].width = 4.17
            ws.column_dimensions["B"].width = 49.17
            ws.column_dimensions["C"].width = 49.17
            ws.column_dimensions["D"].width = 49.17
            ws.column_dimensions["E"].width = 49.17

            # Borders
            for i in range(1, 6):
                ws.cell(row = row, column = i).border = Border(top = Side(style = "medium"), bottom = Side(style = "thin"))
                ws.cell(row = row, column = i).font = Font(bold = True)
            for i in range(1, 6):
                ws.cell(row = row + len(file.translations), column = i).border = Border(bottom = Side(style = "medium"))

            # Alignment ant wrap text
            for row in ws.iter_rows(min_row = row):
                for cell in row:
                    cell.alignment = Alignment(horizontal = 'left',
                                                vertical = 'top',
                                                wrapText = True)
